from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from .models import Produkt, Sprzedaz, Zadanie, Ekspozycja, GrupaProduktowa, Marka, KlientCounter, DzienPracy, GrafikPracy, PunktChecklisty, OdpowiedzChecklisty, TopGroup, TopSubGroup, TopListEntry
from .forms import ZadanieForm, ProduktForm
import openpyxl
from openpyxl.utils.exceptions import InvalidFileException
from zipfile import BadZipFile
from decimal import Decimal, InvalidOperation
import logging
from django.utils import timezone
from django.db.models import Sum, Count, F
from collections import defaultdict
from datetime import timedelta, datetime
from rapidfuzz import fuzz, process 
import re
from django.db import models
from django.http import HttpResponse, JsonResponse
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.models import User
from django.contrib.auth import logout

def custom_logout(request):
    logout(request)
    return redirect('login')
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import calendar

@login_required
def test_template(request):
    try:
        template = get_template('produkty/import_form.html')
        return render(request, 'produkty/import_form.html')  # Jeśli działa, to powinno pokazać poprawny szablon
    except TemplateDoesNotExist:
        return render(request, 'produkty/home.html', {'error': 'Szablon nie istnieje'})

@login_required
def home(request):
    today = datetime.now().date()
    # Tydzień zaczyna się w poniedziałek (weekday() == 0)
    start_of_week = today - timedelta(days=today.weekday())
    end_of_week = start_of_week + timedelta(days=6)
    
    start_of_month = today.replace(day=1)
    # Znajdź ostatni dzień miesiąca
    end_of_month = (start_of_month + timedelta(days=32)).replace(day=1) - timedelta(days=1)

    sprzedaz_tygodniowa = Sprzedaz.objects.filter(data_sprzedazy__range=[start_of_week, end_of_week])
    sprzedaz_miesieczna = Sprzedaz.objects.filter(data_sprzedazy__range=[start_of_month, end_of_month])

    # Obliczanie sum prowizji
    tygodniowa_suma = sprzedaz_tygodniowa.aggregate(
        total=Sum(F('liczba_sztuk') * F('produkt__stawka') + F('prowizja'), output_field=models.DecimalField())
    )['total'] or Decimal('0.00')

    miesieczna_suma = sprzedaz_miesieczna.aggregate(
        total=Sum(F('liczba_sztuk') * F('produkt__stawka') + F('prowizja'), output_field=models.DecimalField())
    )['total'] or Decimal('0.00')


    # Pobierz licznik klientów na dziś
    licznik, _ = KlientCounter.objects.get_or_create(data=today)

    najczesciej_sprzedawane = Sprzedaz.objects.filter(
        data_sprzedazy__gte=start_of_month
    ).values('produkt__model').annotate(
        liczba_sztuk=Sum('liczba_sztuk')
    ).order_by('-liczba_sztuk')[:5]

    context = {
        'najczesciej_sprzedawane': najczesciej_sprzedawane,
        'tygodniowa_suma': tygodniowa_suma,
        'miesieczna_suma': miesieczna_suma,
        'licznik_klientow': licznik.liczba_klientow,
    }

    return render(request, 'produkty/home.html', context)
# Ustaw logger
logger = logging.getLogger(__name__)

@login_required
def import_excel(request):
    if request.method == 'POST' and 'file' in request.FILES:
        excel_file = request.FILES['file']
        try:
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active
        except (InvalidFileException, BadZipFile) as e:
            logger.error(f"Nieprawidłowy plik Excel: {e}")
            return render(request, 'produkty/import_form.html', {
                'error': 'Nieprawidłowy plik Excel. Upewnij się, że przesyłasz prawidłowy plik.'
            })
        except Exception as e:
            logger.error(f"Nieoczekiwany błąd podczas wczytywania pliku: {e}")
            return render(request, 'produkty/import_form.html', {
                'error': 'Wystąpił błąd podczas przetwarzania pliku.'
            })

        existing_products = {p.model: p for p in Produkt.objects.all()}
        products_to_create = []
        products_to_update = []
        models_seen = set()

        for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if len(row) < 4:
                logger.warning(f"Wiersz {idx} pominięty - niewystarczająca liczba kolumn.")
                continue

            grupa_towarowa, marka, model, stawka = row[0], row[1], row[2], row[3]

            if not model:
                logger.warning(f"Wiersz {idx} pominięty - brak modelu.")
                continue

            model_str = str(model).strip()
            if model_str in models_seen:
                continue
            models_seen.add(model_str)

            try:
                if stawka is None or stawka == '':
                    stawka_val = Decimal(0)
                else:
                    stawka_val = Decimal(str(stawka).replace(',', '.'))
            except (TypeError, InvalidOperation, ValueError) as e:
                logger.error(f"Błąd konwersji stawki w wierszu {idx}: {stawka}. Ustawiono 0. Błąd: {e}")
                stawka_val = Decimal(0)

            grupa_val = str(grupa_towarowa).strip() if grupa_towarowa else 'Nieznana'
            marka_val = str(marka).strip() if marka else 'Nieznana'

            if model_str in existing_products:
                p = existing_products[model_str]
                if p.stawka != stawka_val or p.grupa_towarowa != grupa_val or p.marka != marka_val:
                    p.stawka = stawka_val
                    p.grupa_towarowa = grupa_val
                    p.marka = marka_val
                    products_to_update.append(p)
            else:
                products_to_create.append(Produkt(
                    model=model_str,
                    stawka=stawka_val,
                    grupa_towarowa=grupa_val,
                    marka=marka_val
                ))

        if products_to_create:
            Produkt.objects.bulk_create(products_to_create, batch_size=500)
        
        if products_to_update:
            Produkt.objects.bulk_update(products_to_update, ['stawka', 'grupa_towarowa', 'marka'], batch_size=500)

        return render(request, 'produkty/import_success.html')
    elif request.method == 'POST':
        # Brak pliku w żądaniu POST
        return render(request, 'produkty/import_form.html', {'error': 'Nie wybrano pliku do importu.'})

    return render(request, 'produkty/import_form.html')

logger = logging.getLogger(__name__)


def _zapisz_sprzedaz_i_zadania(zatwierdzone_modele, data_sprzedazy, nowe_modele_marki=None):
    if nowe_modele_marki is None:
        nowe_modele_marki = {}

    for model in zatwierdzone_modele:
        marka_val = nowe_modele_marki.get(model, 'Nieznana')
        produkt, created = Produkt.objects.get_or_create(
            model=model,
            defaults={'stawka': 0, 'grupa_towarowa': 'NIEZNANA', 'marka': marka_val}
        )
        if not created and produkt.marka == 'Nieznana' and marka_val != 'Nieznana':
            produkt.marka = marka_val
            produkt.save()
        sprzedaz = Sprzedaz.objects.create(
            produkt=produkt,
            data_sprzedazy=data_sprzedazy,
            liczba_sztuk=1
        )

        wszystkie_zadania = Zadanie.objects.filter(data_start__lte=data_sprzedazy, data_koniec__gte=data_sprzedazy)
        for zadanie in wszystkie_zadania:
            if produkt in zadanie.produkty.all():
                sprzedane = Sprzedaz.objects.filter(
                    produkt__in=zadanie.produkty.all(),
                    data_sprzedazy__range=[zadanie.data_start, zadanie.data_koniec]
                ).aggregate(models.Sum('liczba_sztuk'))['liczba_sztuk__sum'] or 0

                if zadanie.typ == 'MIX_MNOZNIK' and sprzedane >= zadanie.prog_mix:
                    sprzedaz.prowizja = produkt.stawka * zadanie.mnoznik_mix
                    sprzedaz.save()


@login_required
def sprzedaz(request):
    if request.method == 'POST':
        if 'sugestie_zatwierdzone' in request.POST:
            data_sprzedazy = request.POST.get('data_sprzedazy')

            zatwierdzone_modele = []
            for key in request.POST:
                if key.startswith('model_'):
                    model = request.POST[key].strip().upper()
                    zatwierdzone_modele.append(model)

            nowe_modele_marki = {}
            for key in request.POST:
                if key.startswith('nowy_model_'):
                    model_idx = key.split('_')[-1]
                    model = request.POST[key].strip().upper()
                    marka_nazwa = request.POST.get(f'marka_dla_{model_idx}')
                    if not marka_nazwa:
                        marka_nazwa = 'Nieznana'
                    nowe_modele_marki[model] = marka_nazwa
                    zatwierdzone_modele.append(model)

            _zapisz_sprzedaz_i_zadania(zatwierdzone_modele, data_sprzedazy, nowe_modele_marki)
            return redirect('produkty:sprzedaz_sukces')

        data_sprzedazy = request.POST.get('data_sprzedazy')
        modele_sprzedazy = request.POST.get('modele_sprzedazy')

        if modele_sprzedazy:
            modele = [model.strip().upper() for model in modele_sprzedazy.split('\n') if model.strip()]
        else:
            modele = []

        wszystkie_modele = Produkt.objects.values_list('model', flat=True)
        marki = Produkt.objects.exclude(marka__isnull=True).exclude(marka='').exclude(marka='Nieznana').exclude(marka='NIEZNANA').values_list('marka', flat=True).distinct().order_by('marka')

        sugestie = []
        zatwierdzone_modele = []
        nieznane_modele = []

        for model in modele:
            if model in wszystkie_modele:
                zatwierdzone_modele.append(model)
            else:
                najlepszy_wynik = process.extractOne(model, wszystkie_modele, scorer=fuzz.ratio)
                if najlepszy_wynik and najlepszy_wynik[1] > 80:
                    sugestie.append((model, najlepszy_wynik[0]))
                else:
                    nieznane_modele.append(model)

        if sugestie or nieznane_modele:
            context = {
                'sugestie': sugestie,
                'nieznane_modele': nieznane_modele,
                'data_sprzedazy': data_sprzedazy,
                'zatwierdzone_modele': zatwierdzone_modele,
                'marki': marki,
            }
            return render(request, 'produkty/sprzedaz_sugestie.html', context)

        _zapisz_sprzedaz_i_zadania(zatwierdzone_modele, data_sprzedazy)
        return redirect('produkty:sprzedaz_sukces')

    prefilled_model = request.GET.get('model', '')
    today_str = timezone.now().date().strftime('%Y-%m-%d')
    context = {
        'prefilled_model': prefilled_model,
        'today': today_str,
    }
    return render(request, 'produkty/sprzedaz.html', context)

@login_required
def sprzedaz_sukces(request):
    return render(request, 'produkty/sprzedaz_sukces.html')

@login_required
def podsumowanie_sprzedazy(request):
    data_od_str = request.GET.get('data_od')
    data_do_str = request.GET.get('data_do')
    produkt_nazwa = request.GET.get('produkt')
    marka_nazwa = request.GET.get('marka')

    # Normalize empty strings to None
    if not data_od_str:
        data_od_str = None
    if not data_do_str:
        data_do_str = None

    if not data_od_str and not data_do_str:
        today = datetime.now().date()
        start_of_month = today.replace(day=1)
        end_of_month = (start_of_month + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        
        data_od = start_of_month
        data_do = end_of_month
    else:
        try:
            data_od = datetime.strptime(data_od_str, '%Y-%m-%d').date() if data_od_str else None
        except ValueError:
            data_od = None
        try:
            data_do = datetime.strptime(data_do_str, '%Y-%m-%d').date() if data_do_str else None
        except ValueError:
            data_do = None

    # Base monthly navigation on data_od if available, otherwise today
    base_date = data_od if data_od else datetime.now().date()
    current_month_start = base_date.replace(day=1)
    
    prev_month_end = current_month_start - timedelta(days=1)
    prev_month_start = prev_month_end.replace(day=1)

    next_month_start = (current_month_start + timedelta(days=32)).replace(day=1)
    next_month_end = (next_month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)

    sprzedaz_qs = Sprzedaz.objects.select_related('produkt').all()

    if data_od:
        sprzedaz_qs = sprzedaz_qs.filter(data_sprzedazy__gte=data_od)
    if data_do:
        sprzedaz_qs = sprzedaz_qs.filter(data_sprzedazy__lte=data_do)
    if produkt_nazwa:
        sprzedaz_qs = sprzedaz_qs.filter(produkt__model__icontains=produkt_nazwa)
    if marka_nazwa:
        sprzedaz_qs = sprzedaz_qs.filter(produkt__marka__icontains=marka_nazwa)

    # Annotate with the commission for each sale
    sprzedaz_annotated = sprzedaz_qs.annotate(
        obliczona_prowizja=models.F('liczba_sztuk') * models.F('produkt__stawka')
    )

    sprzedaz_podsumowanie = (
        sprzedaz_annotated
        .values('produkt__marka', 'produkt__model', 'produkt__stawka')
        .annotate(
            liczba_sztuk=Sum('liczba_sztuk'),
            suma_prowizji=Sum('obliczona_prowizja')
        )
        .order_by('-suma_prowizji')
    )

    # Dictionary for tests
    sprzedaz_dict = {}
    for s in sprzedaz_podsumowanie:
        marka = s['produkt__marka'] or ''
        model = s['produkt__model'] or ''
        key = f"{marka}_{model}"
        sprzedaz_dict[key] = {
            'marka': marka,
            'model': model,
            'stawka': s['produkt__stawka'],
            'suma_prowizji': s['suma_prowizji'],
            'liczba_sztuk': s['liczba_sztuk']
        }

    # Task rewards computation
    task_rewards = []
    if data_od and data_do:
        tasks = Zadanie.objects.filter(data_start__lte=data_do, data_koniec__gte=data_od)
        for task in tasks:
            sales_in_task = Sprzedaz.objects.filter(
                produkt__in=task.produkty.all(),
                data_sprzedazy__range=(task.data_start, task.data_koniec)
            )
            total_sold = sales_in_task.aggregate(suma=Sum('liczba_sztuk'))['suma'] or 0
            
            premia = Decimal("0.00")
            if task.prog_2 and total_sold >= task.prog_2:
                premia = task.prog_2_premia or Decimal("0.00")
            elif task.prog_1 and total_sold >= task.prog_1:
                premia = task.prog_1_premia or Decimal("0.00")
                
            if premia > 0:
                task_rewards.append({
                    'nazwa': task.nazwa,
                    'sprzedane': total_sold,
                    'premia': premia
                })

    grupa_beko_marki = ['BEKO', 'GRUNDIG']
    grupa_whirlpool_marki = ['WHIRLPOOL', 'INDESIT', 'HOTPOINT']

    grupa_beko = []
    grupa_whirlpool = []
    inne = []

    for s in sprzedaz_podsumowanie:
        marka = s['produkt__marka'].upper() if s['produkt__marka'] else ''
        if marka in grupa_beko_marki:
            grupa_beko.append(s)
        elif marka in grupa_whirlpool_marki:
            grupa_whirlpool.append(s)
        else:
            inne.append(s)

    agregaty = sprzedaz_annotated.aggregate(
        calkowita_liczba_sztuk=Sum('liczba_sztuk'),
        calkowita_prowizja=Sum('obliczona_prowizja')
    )

    # Pobranie unikalnych marek dla filtra
    dostepne_marki = Produkt.objects.exclude(marka__isnull=True).exclude(marka='').exclude(marka='Nieznana').exclude(marka='NIEZNANA').values_list('marka', flat=True).distinct().order_by('marka')

    context = {
        'grupa_beko': grupa_beko,
        'grupa_whirlpool': grupa_whirlpool,
        'grupa_inne': inne,
        'liczba_sztuk': agregaty['calkowita_liczba_sztuk'] or 0,
        'calkowita_prowizja': agregaty['calkowita_prowizja'] or 0,
        'data_od': data_od,
        'data_do': data_do,
        'produkt': produkt_nazwa or '',
        'marka': marka_nazwa or '',
        'prev_month_data_od': prev_month_start.strftime('%Y-%m-%d'),
        'prev_month_data_do': prev_month_end.strftime('%Y-%m-%d'),
        'next_month_data_od': next_month_start.strftime('%Y-%m-%d'),
        'next_month_data_do': next_month_end.strftime('%Y-%m-%d'),
        'sprzedaz': sprzedaz_dict,
        'task_rewards': task_rewards,
        'dostepne_marki': dostepne_marki,
    }

    return render(request, 'produkty/podsumowanie_sprzedazy.html', context)

@login_required
def wyciagnij_liste_modeli(request):
    if request.method == 'POST':
        tekst = request.POST.get('tekst_sprzedazy', '')
        # Wyciągnięcie sekcji MOJE
        wzorzec_moje = re.compile(r'MOJE\s*(.*?)\s*(INNE|$)', re.S)
        wynik_moje = wzorzec_moje.search(tekst)
        if wynik_moje:
            lista_modeli = wynik_moje.group(1).split('\n')
            # Filtrowanie tylko rzeczywistych modeli (np. brak spacji, tylko duże litery i cyfry)
            lista_modeli = [
                model.strip() for model in lista_modeli
                if re.match(r'^[A-Z0-9]+$', model.strip())
            ]
        else:
            lista_modeli = []

        context = {'lista_modeli': lista_modeli}
        return render(request, 'produkty/wyciagnij_liste_modeli.html', context)
    else:
        return render(request, 'produkty/wyciagnij_liste_modeli.html')



@login_required
def ekspozycja_form(request, grupa_id):
    grupa = get_object_or_404(GrupaProduktowa, id=grupa_id)
    marki = Marka.objects.all()

    # Pobierz wszystkie grupy w kolejności po ID
    wszystkie_grupy = list(GrupaProduktowa.objects.order_by('id'))
    obecny_index = wszystkie_grupy.index(grupa)

    # Określ poprzednią i następną grupę
    poprzednia_grupa = wszystkie_grupy[obecny_index - 1] if obecny_index > 0 else None
    nastepna_grupa = wszystkie_grupy[obecny_index + 1] if obecny_index < len(wszystkie_grupy) - 1 else None

    if request.method == 'POST':
        # Usuń istniejące rekordy dla tej grupy, aby uniknąć duplikatów
        Ekspozycja.objects.filter(grupa=grupa).delete()

        for marka in marki:
            liczba = int(request.POST.get(f'marka_{marka.id}', 0))
            Ekspozycja.objects.create(
                grupa=grupa,
                marka=marka,
                liczba=liczba
            )
        
        # Przejdź do następnej grupy lub wyświetl podsumowanie
        if nastepna_grupa:
            return redirect('produkty:ekspozycja_form', grupa_id=nastepna_grupa.id)
        return redirect('produkty:ekspozycja_summary')

    return render(request, 'produkty/ekspozycja_form.html', {
        'grupa': grupa,
        'marki': marki,
        'poprzednia_grupa': poprzednia_grupa,
    })

        
@login_required
def ekspozycja_summary(request):
    # Pobierz wszystkie dane ekspozycji z bazy danych
    ekspozycje = Ekspozycja.objects.select_related('grupa', 'marka').all()

    # Przygotuj dane do wyświetlenia w podsumowaniu
    grupy = GrupaProduktowa.objects.all()

    dane_podsumowania = []
    for grupa in grupy:
        marki_ekspozycji = ekspozycje.filter(grupa=grupa)
        dane_podsumowania.append({
            'grupa': grupa.nazwa,
            'marki': marki_ekspozycji,
        })

    return render(request, 'produkty/ekspozycja_summary.html', {
        'dane_podsumowania': dane_podsumowania,
    })

@login_required
def reset_sprzedaz(request):
    if request.method == 'POST':
        Sprzedaz.objects.all().delete()
        return redirect('produkty:podsumowanie_sprzedazy')
    return redirect('produkty:podsumowanie_sprzedazy')

@login_required
def klienci(request):
    today = timezone.now().date()
    counter, created = KlientCounter.objects.get_or_create(data=today)
    
    context = {
        'counter': counter,
    }
    return render(request, 'produkty/klienci.html', context)

@login_required
def zmien_licznik(request, operacja):
    if request.method == 'POST':
        today = timezone.now().date()
        counter, created = KlientCounter.objects.get_or_create(data=today)
        
        if operacja == 'plus':
            counter.liczba_klientow += 1
        elif operacja == 'minus' and counter.liczba_klientow > 0:
            counter.liczba_klientow -= 1
            
        counter.save()
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'liczba_klientow': counter.liczba_klientow})
        
    return redirect('produkty:klienci')

@login_required
def eksportuj_ekspozycje_xlsx(request):
    # Tworzenie nowego arkusza Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ekspozycja"
    
    # Pobieranie danych
    grupy = GrupaProduktowa.objects.all().order_by('id')
    marki = Marka.objects.all().order_by('id')
    ekspozycje = Ekspozycja.objects.select_related('grupa', 'marka').all()
    
    # Stylizacja
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Nagłówki
    ws['A1'] = "Lista TOTAL duża AGD udziały producentów"
    ws['A1'].font = Font(bold=True)
    ws.merge_cells('A1:BV1')
    
    ws['A2'] = "Kategoria"
    ws['A2'].fill = header_fill
    ws['A2'].font = header_font
    ws['A2'].alignment = header_alignment
    ws['A2'].border = thin_border
    
    ws['B2'] = "Produkty TOTAL suma"
    ws['B2'].fill = header_fill
    ws['B2'].font = header_font
    ws['B2'].alignment = header_alignment
    ws['B2'].border = thin_border
    
    # Dodawanie nagłówków marek
    col_index = 3
    for marka in marki:
        col_letter = get_column_letter(col_index)
        ws[f'{col_letter}2'] = f"{marka.nazwa} ilość"
        ws[f'{col_letter}2'].fill = header_fill
        ws[f'{col_letter}2'].font = header_font
        ws[f'{col_letter}2'].alignment = header_alignment
        ws[f'{col_letter}2'].border = thin_border
        
        col_letter = get_column_letter(col_index + 1)
        ws[f'{col_letter}2'] = f"{marka.nazwa} %"
        ws[f'{col_letter}2'].fill = header_fill
        ws[f'{col_letter}2'].font = header_font
        ws[f'{col_letter}2'].alignment = header_alignment
        ws[f'{col_letter}2'].border = thin_border
        
        col_index += 2
    
    # Wypełnianie danymi
    row_index = 3
    for grupa in grupy:
        # Nazwa grupy
        ws[f'A{row_index}'] = grupa.nazwa
        ws[f'A{row_index}'].border = thin_border
        
        # Obliczanie sumy dla grupy
        suma_grupa = sum([ekspozycja.liczba for ekspozycja in ekspozycje.filter(grupa=grupa)])
        ws[f'B{row_index}'] = suma_grupa
        ws[f'B{row_index}'].border = thin_border
        
        # Wypełnianie danymi dla każdej marki
        col_index = 3
        for marka in marki:
            try:
                ekspozycja = ekspozycje.get(grupa=grupa, marka=marka)
                liczba = ekspozycja.liczba
            except Ekspozycja.DoesNotExist:
                liczba = 0
            
            # Ilość
            col_letter = get_column_letter(col_index)
            ws[f'{col_letter}{row_index}'] = liczba
            ws[f'{col_letter}{row_index}'].border = thin_border
            
            # Procent
            col_letter = get_column_letter(col_index + 1)
            if suma_grupa > 0:
                procent = (liczba / suma_grupa) * 100
                ws[f'{col_letter}{row_index}'] = f"{procent:.1f}%"
            else:
                ws[f'{col_letter}{row_index}'] = "#DIV/0!"
            ws[f'{col_letter}{row_index}'].border = thin_border
            
            col_index += 2
        
        row_index += 1
    
    # Dodawanie wiersza podsumowania
    ws[f'A{row_index}'] = "Udziały producenta w ekspozycji produktów TOTAL"
    ws[f'A{row_index}'].border = thin_border
    
    # Obliczanie sumy całkowitej
    suma_total = sum([ekspozycja.liczba for ekspozycja in ekspozycje])
    ws[f'B{row_index}'] = suma_total
    ws[f'B{row_index}'].border = thin_border
    
    # Wypełnianie podsumowania dla każdej marki
    col_index = 3
    for marka in marki:
        suma_marka = sum([ekspozycja.liczba for ekspozycja in ekspozycje.filter(marka=marka)])
        
        # Ilość
        col_letter = get_column_letter(col_index)
        ws[f'{col_letter}{row_index}'] = suma_marka
        ws[f'{col_letter}{row_index}'].border = thin_border
        
        # Procent
        col_letter = get_column_letter(col_index + 1)
        if suma_total > 0:
            procent = (suma_marka / suma_total) * 100
            ws[f'{col_letter}{row_index}'] = f"{procent:.1f}%"
        else:
            ws[f'{col_letter}{row_index}'] = "#DIV/0!"
        ws[f'{col_letter}{row_index}'].border = thin_border
        
        col_index += 2
    
    # Dostosowanie szerokości kolumn
    for col in range(1, col_index):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Zapisywanie do bufora i zwracanie jako odpowiedź HTTP
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=ekspozycja.xlsx'
    
    return response

@login_required
def delete_all_models(request):
    if request.method == 'POST':
        # Usuwanie wszystkich produktów
        Produkt.objects.all().delete()
        return render(request, 'produkty/import_success.html', {'message': 'Wszystkie modele zostały usunięte.'})
    
    # Jeśli nie jest to POST, przekieruj na stronę importu
    return redirect('produkty:import_excel')

@login_required
def zadania_management(request):
    """Widok do zarządzania zadaniami"""
    today = datetime.now().date()
    return redirect('produkty:zadania_view', year=today.year, month=today.month)

@login_required
def zadania_view(request, year, month):
    """Widok do wyświetlania zadań w ujęciu miesięcznym"""
    zadania = Zadanie.objects.filter(
        data_start__year=year, data_start__month=month
    ).order_by('-data_start')

    first_day_of_month = datetime(year, month, 1)
    prev_month_date = first_day_of_month - timedelta(days=1)
    next_month_date = (first_day_of_month + timedelta(days=32)).replace(day=1)

    context = {
        'zadania': zadania,
        'year': year,
        'month': month,
        'month_name': calendar.month_name[month],
        'prev_year': prev_month_date.year,
        'prev_month': prev_month_date.month,
        'next_year': next_month_date.year,
        'next_month': next_month_date.month,
    }
    return render(request, 'produkty/zadania.html', context)


@login_required
def zadanie_dodaj(request):
    """Widok do dodawania nowego zadania"""
    import json
    from datetime import datetime
    
    if request.method == 'POST':
        json_data = request.POST.get('zadanie_json')
        if json_data and json_data.strip():
            try:
                dane = json.loads(json_data)
                
                # Obsługa zarówno pojedynczego obiektu, jak i listy zadań
                if isinstance(dane, list):
                    zadania_dane = dane
                else:
                    zadania_dane = [dane]

                last_redirect_date = None
                
                for item in zadania_dane:
                    nazwa = item.get('nazwa', 'Zadanie z JSON')
                    opis = item.get('opis', '')
                    data_start = item.get('data_start')
                    data_koniec = item.get('data_koniec')
                    target = item.get('target', 'ilosc')
                    prog_1 = item.get('prog_1')
                    prog_1_premia = item.get('prog_1_premia')
                    prog_2 = item.get('prog_2')
                    prog_2_premia = item.get('prog_2_premia')
                    typ = item.get('typ', 'KONKRETNE_MODELE')
                    mnoznik_mix = item.get('mnoznik_mix')
                    prog_mix = item.get('prog_mix')
                    modele = item.get('modele', [])

                    zadanie = Zadanie.objects.create(
                        nazwa=nazwa, opis=opis,
                        data_start=data_start, data_koniec=data_koniec,
                        target=target,
                        prog_1=prog_1, prog_1_premia=prog_1_premia,
                        prog_2=prog_2, prog_2_premia=prog_2_premia,
                        typ=typ, mnoznik_mix=mnoznik_mix, prog_mix=prog_mix
                    )

                    for model in modele:
                        produkt, _ = Produkt.objects.get_or_create(
                            model=model.strip().upper(),
                            defaults={'stawka': 0, 'grupa_towarowa': 'NIEZNANA', 'marka': 'Nieznana'}
                        )
                        zadanie.produkty.add(produkt)
                    
                    if data_start:
                        try:
                            last_redirect_date = datetime.strptime(data_start, '%Y-%m-%d')
                        except ValueError:
                            pass
                
                if last_redirect_date:
                    return redirect('produkty:zadania_view', year=last_redirect_date.year, month=last_redirect_date.month)
                return redirect('produkty:zadania_management')
            except json.JSONDecodeError as e:
                logger.error(f"Błąd parsowania JSON dla zadania: {e}")
                form = ZadanieForm(request.POST)
                form.add_error(None, "Nieprawidłowy format JSON. Popraw błędy i spróbuj ponownie.")
            except Exception as e:
                logger.error(f"Błąd podczas tworzenia zadania z JSON: {e}")
                form = ZadanieForm(request.POST)
                form.add_error(None, f"Błąd bazy danych podczas dodawania zadania: {e}")
        else:
            form = ZadanieForm(request.POST)
            if form.is_valid():
                zadanie = form.save()
                try:
                    return redirect('produkty:zadania_view', year=zadanie.data_start.year, month=zadanie.data_start.month)
                except Exception:
                    return redirect('produkty:zadania_management')
    else:
        form = ZadanieForm()
    
    return render(request, 'produkty/zadanie_form.html', {
        'form': form,
        'selected_products_pks': []
    })

@login_required
def zadanie_edytuj(request, zadanie_id):
    """Widok do edycji istniejącego zadania"""
    zadanie = get_object_or_404(Zadanie, id=zadanie_id)
    
    if request.method == 'POST':
        form = ZadanieForm(request.POST, instance=zadanie)
        if form.is_valid():
            zadanie = form.save()
            try:
                return redirect('produkty:zadania_view', year=zadanie.data_start.year, month=zadanie.data_start.month)
            except Exception:
                return redirect('produkty:zadania_management')
    else:
        form = ZadanieForm(instance=zadanie)

    selected_products_pks = list(zadanie.produkty.values_list('pk', flat=True))

    return render(request, 'produkty/zadanie_form.html', {
        'form': form, 
        'zadanie': zadanie,
        'selected_products_pks': selected_products_pks
    })

@login_required
def zadanie_usun(request, zadanie_id):
    """Widok do usuwania zadania"""
    zadanie = get_object_or_404(Zadanie, id=zadanie_id)
    
    if request.method == 'POST':
        month = zadanie.data_start.month
        year = zadanie.data_start.year
        zadanie.delete()
        return redirect('produkty:zadania_view', year=year, month=month)
    
    return render(request, 'produkty/zadanie_usun.html', {'zadanie': zadanie})

@login_required
def szczegoly_zadania(request, zadanie_id):
    zadanie = get_object_or_404(Zadanie, id=zadanie_id)
    modele_w_zadaniu = zadanie.produkty.all()

    sprzedaz_w_okresie = Sprzedaz.objects.filter(
        produkt__in=modele_w_zadaniu,
        data_sprzedazy__range=(zadanie.data_start, zadanie.data_koniec)
    )

    postep = sprzedaz_w_okresie.aggregate(suma=Sum('liczba_sztuk'))['suma'] or 0
    prog_1_status = False
    if zadanie.prog_1 and postep >= zadanie.prog_1:
        prog_1_status = True

    prog_2_status = False
    if zadanie.prog_2 and postep >= zadanie.prog_2:
        prog_2_status = True

    sprzedane_modele = sprzedaz_w_okresie.values('produkt__model', 'produkt__marka').annotate(
        ilosc=Sum('liczba_sztuk')
    ).order_by('-ilosc')

    # Analiza historyczna
    days = request.GET.get('days', '90')
    if days not in ['30', '60', '90']:
        days = '90'
    days_int = int(days)
    
    from datetime import datetime, timedelta
    today_date = datetime.now().date()
    start_date_hist = today_date - timedelta(days=days_int)
    
    historyczne_sprzedaze = Sprzedaz.objects.filter(
        produkt__in=modele_w_zadaniu,
        data_sprzedazy__range=(start_date_hist, today_date)
    ).values('produkt').annotate(ilosc=Sum('liczba_sztuk'))
    
    historia_dict = {item['produkt']: item['ilosc'] for item in historyczne_sprzedaze}
    
    analiza_dane = []
    for prod in modele_w_zadaniu:
        sztuk = historia_dict.get(prod.id, 0)
        analiza_dane.append({
            'produkt': prod,
            'sztuk': sztuk
        })
    analiza_dane.sort(key=lambda x: x['sztuk'], reverse=True)

    context = {
        'zadanie': zadanie,
        'postep': postep,
        'sprzedane_modele': sprzedane_modele,
        'modele_w_zadaniu': modele_w_zadaniu,
        'prog_1_status': prog_1_status,
        'prog_2_status': prog_2_status,
        'analiza_dane': analiza_dane,
        'selected_days': days_int,
    }

    return render(request, 'produkty/szczegoly_zadania.html', context)

@login_required
def otwarcie_dnia(request):
    import datetime
    today = datetime.date.today()
    from django.utils import timezone
    from django.contrib import messages
    
    dzien, created = DzienPracy.objects.get_or_create(
        user=request.user,
        data=today,
        defaults={'czas_otwarcia': timezone.now()}
    )
    if created:
        messages.success(request, "Pomyślnie rozpoczęto dzień pracy!")
    else:
        messages.info(request, "Dzień pracy został już rozpoczęty wcześniej.")
    return redirect('produkty:home')

@login_required
def zamkniecie_dnia(request):
    import datetime
    from django.utils import timezone
    from django.contrib import messages
    today = datetime.date.today()
    
    dzien = DzienPracy.objects.filter(user=request.user, data=today).first()
    if not dzien:
        return redirect('produkty:otwarcie_dnia')
        
    # Unikamy dublowania punktów o tej samej treści w bazie danych
    all_punkty = list(PunktChecklisty.objects.filter(aktywny=True))
    seen_texts = set()
    punkty = []
    for p in all_punkty:
        if p.tekst not in seen_texts:
            seen_texts.add(p.tekst)
            punkty.append(p)
    
    if request.method == 'POST':
        notatka = request.POST.get('notatka_sprzedaz', '')
        dzien.notatka_sprzedaz = notatka
        dzien.czas_zamkniecia = timezone.now()
        dzien.save()
        
        for punkt in punkty:
            wykonano = request.POST.get(f'punkt_{punkt.id}') == 'on'
            OdpowiedzChecklisty.objects.update_or_create(
                dzien_pracy=dzien,
                punkt=punkt,
                defaults={'wykonano': wykonano}
            )
            
        messages.success(request, "Dzień pracy został pomyślnie zakończony, a raport zapisany!")
        return redirect('produkty:home')
        
    odpowiedzi = OdpowiedzChecklisty.objects.filter(dzien_pracy=dzien)
    odp_dict = {o.punkt_id: o.wykonano for o in odpowiedzi}
    
    # Obliczanie postępu w zadaniach
    zadania_progress = []
    dzisiejsze_sprzedaze = Sprzedaz.objects.filter(data_sprzedazy=today)
    
    aktywne_zadania = Zadanie.objects.filter(data_start__lte=today, data_koniec__gte=today)
    for zadanie in aktywne_zadania:
        # Zlicz ile sztuk z produktów w zadaniu sprzedano dzisiaj
        sprzedane_dzis = dzisiejsze_sprzedaze.filter(produkt__in=zadanie.produkty.all()).aggregate(models.Sum('liczba_sztuk'))['liczba_sztuk__sum'] or 0
        if sprzedane_dzis > 0:
            zadania_progress.append({
                'zadanie': zadanie.nazwa,
                'sztuki': sprzedane_dzis
            })
    
    context = {
        'dzien': dzien,
        'punkty': punkty,
        'odp_dict': odp_dict,
        'zadania_progress': zadania_progress
    }
    return render(request, 'produkty/zamkniecie_dnia.html', context)

@login_required
def checklista_podglad(request):
    import datetime
    today = datetime.date.today()
    dzien = DzienPracy.objects.filter(user=request.user, data=today).first()
    
    if not dzien:
        messages.warning(request, "Musisz rozpocząć dzień pracy, aby podejrzeć checklistę.")
        return redirect('produkty:home')
        
    all_punkty = list(PunktChecklisty.objects.filter(aktywny=True))
    seen_texts = set()
    punkty = []
    for p in all_punkty:
        if p.tekst not in seen_texts:
            seen_texts.add(p.tekst)
            punkty.append(p)
            
    odpowiedzi = OdpowiedzChecklisty.objects.filter(dzien_pracy=dzien)
    odp_dict = {o.punkt_id: o.wykonano for o in odpowiedzi}
    
    if request.method == 'POST':
        for punkt in punkty:
            wykonano = request.POST.get(f'punkt_{punkt.id}') == 'on'
            OdpowiedzChecklisty.objects.update_or_create(
                dzien_pracy=dzien,
                punkt=punkt,
                defaults={'wykonano': wykonano}
            )
        messages.success(request, "Checklista została zaktualizowana.")
        return redirect('produkty:checklista_podglad')

    return render(request, 'produkty/checklista_podglad.html', {
        'dzien': dzien,
        'punkty': punkty,
        'odp_dict': odp_dict
    })

@login_required
def grafik_view(request):
    import datetime
    today = datetime.date.today()
    month = int(request.GET.get('month', today.month))
    year = int(request.GET.get('year', today.year))
    
    is_admin = request.user.is_superuser or request.user.is_staff
    
    if is_admin:
        wybrany_user = request.GET.get('user', request.user.id)
        try:
            wybrany_user_id = int(wybrany_user) if wybrany_user else request.user.id
        except ValueError:
            wybrany_user_id = request.user.id
        grafik = GrafikPracy.objects.filter(data__year=year, data__month=month, user_id=wybrany_user_id).order_by('data')
        users = User.objects.all()
    else:
        grafik = GrafikPracy.objects.filter(user=request.user, data__year=year, data__month=month).order_by('data')
        users = None
        wybrany_user_id = request.user.id

    context = {
        'grafik': grafik,
        'year': year,
        'month': month,
        'is_admin': is_admin,
        'users': users,
        'wybrany_user': wybrany_user_id
    }
    return render(request, 'produkty/grafik.html', context)

@login_required
def import_grafiku(request):
    if request.method == 'POST' and 'file' in request.FILES:
        excel_file = request.FILES['file']
        user_id = request.POST.get('user_id', request.user.id)
        
        if not (request.user.is_superuser or request.user.is_staff):
            user_id = request.user.id
            
        try:
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            sheet = wb.active
        except Exception as e:
            logger.error(f"Error loading excel: {e}")
            users = User.objects.all() if (request.user.is_superuser or request.user.is_staff) else None
            return render(request, 'produkty/import_grafiku.html', {'error': 'Nieprawidłowy plik.', 'users': users})
            
        import datetime as dt_mod
        from datetime import datetime, time
        from django.contrib.auth.models import User
        user_obj = User.objects.get(id=user_id)
        
        def parse_date(val):
            if isinstance(val, datetime):
                return val.date()
            if isinstance(val, dt_mod.date):
                return val
            if isinstance(val, str) and val.strip():
                clean_val = val.strip()
                for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%Y/%m/%d'):
                    try:
                        return datetime.strptime(clean_val, fmt).date()
                    except ValueError:
                        continue
            return None

        def parse_time(val):
            if isinstance(val, time): 
                return val
            if isinstance(val, datetime): 
                return val.time()
            if isinstance(val, (int, float)) and 0 <= val <= 1:
                total_seconds = int(round(val * 86400))
                hours = (total_seconds // 3600) % 24
                minutes = (total_seconds % 3600) // 60
                return time(hours, minutes)
            if isinstance(val, str) and val.strip():
                val_str = val.strip()
                for fmt in ('%H:%M:%S', '%H:%M', '%I:%M %p', '%I:%M%p'):
                    try: 
                        return datetime.strptime(val_str, fmt).time()
                    except ValueError: 
                        continue
            return None

        # Dynamically detect which column contains the date
        date_col = -1
        for row in sheet.iter_rows(min_row=1, max_row=15, values_only=True):
            if not row:
                continue
            for idx in range(min(len(row), 3)): # Check columns A, B, C (0, 1, 2)
                if parse_date(row[idx]) is not None:
                    date_col = idx
                    break
            if date_col != -1:
                break

        if date_col == -1:
            date_col = 1 # fallback to column B

        for row in sheet.iter_rows(min_row=1, values_only=True):
            if len(row) <= date_col:
                continue
                
            data_val = parse_date(row[date_col])
            if not data_val:
                continue
                
            start_val = row[date_col + 1] if len(row) > date_col + 1 else None
            koniec_val = row[date_col + 2] if len(row) > date_col + 2 else None
            suma_val = row[date_col + 3] if len(row) > date_col + 3 else None
            notatka_val = row[date_col + 4] if len(row) > date_col + 4 else ''
            
            start_time = parse_time(start_val)
            koniec_time = parse_time(koniec_val)
            
            suma_dec = None
            if suma_val:
                try:
                    suma_dec = Decimal(str(suma_val).replace(',', '.'))
                except:
                    pass
                    
            GrafikPracy.objects.update_or_create(
                user=user_obj,
                data=data_val,
                defaults={
                    'godzina_rozpoczecia': start_time,
                    'godzina_zakonczenia': koniec_time,
                    'suma_godzin': suma_dec,
                    'notatka': str(notatka_val) if notatka_val else ''
                }
            )
            
        return redirect('produkty:grafik')
    
    from django.contrib.auth.models import User
    users = User.objects.all() if (request.user.is_superuser or request.user.is_staff) else None
    return render(request, 'produkty/import_grafiku.html', {'users': users})

@login_required
def calendar_view(request, year=None, month=None):
    today = datetime.now().date()
    if year is None or month is None:
        year = today.year
        month = today.month

    # Pobieranie sprzedaży dla danego miesiąca i obliczanie prowizji
    sales_in_month = Sprzedaz.objects.filter(
        data_sprzedazy__year=year,
        data_sprzedazy__month=month
    ).annotate(
        commission=models.F('liczba_sztuk') * models.F('produkt__stawka') + models.F('prowizja')
    ).values('data_sprzedazy').annotate(
        total_items=Sum('liczba_sztuk'),
        total_commission=Sum('commission')
    ).order_by('data_sprzedazy')

    sales_by_day = {
        sale['data_sprzedazy'].day: {
            'total_items': sale['total_items'],
            'total_commission': sale['total_commission']
        } for sale in sales_in_month
    }

    # Ustawienia kalendarza
    cal = calendar.Calendar()
    month_days = cal.monthdayscalendar(year, month)

    # Poprzedni i następny miesiąc
    first_day_of_month = datetime(year, month, 1)
    prev_month_date = first_day_of_month - timedelta(days=1)
    next_month_date = (first_day_of_month + timedelta(days=32)).replace(day=1)

    context = {
        'year': year,
        'month': month,
        'month_name': calendar.month_name[month],
        'month_days': month_days,
        'sales_by_day': sales_by_day,
        'today': today,
        'prev_year': prev_month_date.year,
        'prev_month': prev_month_date.month,
        'next_year': next_month_date.year,
        'next_month': next_month_date.month,
    }
    return render(request, 'produkty/calendar.html', context)



@login_required

def daily_sales_view(request, year, month, day):
    date_of_sales = datetime(year, month, day).date()
    sales = Sprzedaz.objects.filter(data_sprzedazy=date_of_sales).select_related('produkt').annotate(
        total_commission=models.F('liczba_sztuk') * models.F('produkt__stawka') + models.F('prowizja')
    )

    totals = sales.aggregate(
        total_items_sum=Sum('liczba_sztuk'),
        total_commission_sum=Sum('total_commission')
    )

    context = {
        'date_of_sales': date_of_sales,
        'sales': sales,
        'total_items': totals['total_items_sum'],
        'total_commission': totals['total_commission_sum'],
    }
    return render(request, 'produkty/daily_sales.html', context)



@staff_member_required

def delete_sales_for_day(request):

    message = ''

    if request.method == 'POST':

        date_str = request.POST.get('date')

        if date_str:

            try:

                date_to_delete = datetime.strptime(date_str, '%Y-%m-%d').date()

                sales_to_delete = Sprzedaz.objects.filter(data_sprzedazy=date_to_delete)

                count = sales_to_delete.count()

                sales_to_delete.delete()

                message = f'Pomyślnie usunięto {count} rekordów sprzedaży z dnia {date_to_delete}.'

            except ValueError:

                message = 'Nieprawidłowy format daty. Użyj formatu YYYY-MM-DD.'

        else:

            message = 'Proszę wybrać datę.'



    return render(request, 'produkty/delete_sales_for_day.html', {'message': message})

@login_required
def lista_produktow(request):
    produkty = Produkt.objects.all()

    # Filtering
    model_filter = request.GET.get('model', '')
    marka_filter = request.GET.get('marka', '')
    grupa_towarowa_filter = request.GET.get('grupa_towarowa', '')
    prowizja_od = request.GET.get('prowizja_od', '')
    prowizja_do = request.GET.get('prowizja_do', '')

    if model_filter:
        produkty = produkty.filter(model__icontains=model_filter)
    if marka_filter:
        produkty = produkty.filter(marka__icontains=marka_filter)
    if grupa_towarowa_filter:
        produkty = produkty.filter(grupa_towarowa__icontains=grupa_towarowa_filter)
    if prowizja_od:
        try:
            produkty = produkty.filter(stawka__gte=Decimal(prowizja_od))
        except (InvalidOperation, ValueError):
            pass  # Ignore invalid decimal values
    if prowizja_do:
        try:
            produkty = produkty.filter(stawka__lte=Decimal(prowizja_do))
        except (InvalidOperation, ValueError):
            pass  # Ignore invalid decimal values

    # Grouping
    group_by = request.GET.get('group', None)
    grouped_produkty = None

    if group_by in ['marka', 'grupa_towarowa']:
        grouped_produkty = defaultdict(list)
        for produkt in produkty.order_by(group_by):
            grouped_produkty[getattr(produkt, group_by)].append(produkt)
        # Sort the groups by key, handling None values
        grouped_produkty = dict(sorted(grouped_produkty.items(), key=lambda item: (item[0] is None, item[0])))

    # Sorting
    sort_by = request.GET.get('sort', 'model')
    if sort_by not in ['model', 'marka', 'stawka', 'grupa_towarowa']:
        sort_by = 'model'
    
    direction = request.GET.get('dir', 'asc')
    if direction == 'desc':
        sort_by = f'-{sort_by}'
    
    if not grouped_produkty:
        produkty = produkty.order_by(sort_by)

    context = {
        'produkty': produkty,
        'grouped_produkty': grouped_produkty,
        'group_by': group_by,
        'model_filter': model_filter,
        'marka_filter': marka_filter,
        'grupa_towarowa_filter': grupa_towarowa_filter,
        'prowizja_od': prowizja_od,
        'prowizja_do': prowizja_do,
        'sort': request.GET.get('sort', 'model'),
        'dir': request.GET.get('dir', 'asc'),
    }
    return render(request, 'produkty/lista_produktow.html', context)

@login_required
def product_edit(request, product_id):
    produkt = get_object_or_404(Produkt, pk=product_id)
    if request.method == 'POST':
        form = ProduktForm(request.POST, instance=produkt)
        if form.is_valid():
            form.save()
            return redirect('produkty:lista_produktow')
    else:
        form = ProduktForm(instance=produkt)
    return render(request, 'produkty/product_edit.html', {'form': form})

from .models import Funkcja, Podpowiedz

@login_required
def hub_wiedzy(request):
    funkcje = Funkcja.objects.all().order_by('nazwa')
    podpowiedzi = Podpowiedz.objects.all().order_by('tytul')
    return render(request, 'produkty/hub_wiedzy.html', {'funkcje': funkcje, 'podpowiedzi': podpowiedzi})

@login_required
def funkcja_szczegoly(request, funkcja_id):
    funkcja = get_object_or_404(Funkcja, id=funkcja_id)
    produkty = funkcja.produkty.all().order_by('grupa_towarowa', 'model')
    return render(request, 'produkty/funkcja_szczegoly.html', {'funkcja': funkcja, 'produkty': produkty})

@login_required
def podpowiedz_szczegoly(request, podpowiedz_id):
    podpowiedz = get_object_or_404(Podpowiedz, id=podpowiedz_id)
    return render(request, 'produkty/podpowiedz_szczegoly.html', {'podpowiedz': podpowiedz})

@login_required
def historia_sprzedazy_produktu(request, produkt_id):
    produkt = get_object_or_404(Produkt, id=produkt_id)
    sprzedaz_list = Sprzedaz.objects.filter(produkt=produkt).order_by('-data_sprzedazy')
    suma_sztuk = sprzedaz_list.aggregate(total=Sum('liczba_sztuk'))['total'] or 0
    return render(request, 'produkty/historia_sprzedazy_produktu.html', {
        'produkt': produkt,
        'sprzedaz_list': sprzedaz_list,
        'suma_sztuk': suma_sztuk
    })

from .forms import UserCreateForm, UserEditForm

@staff_member_required
def konta_lista(request):
    konta = User.objects.all().order_by('username')
    return render(request, 'produkty/konta_lista.html', {'konta': konta})

@staff_member_required
def konto_dodaj(request):
    if request.method == 'POST':
        form = UserCreateForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Konto zostało utworzone pomyślnie.')
            return redirect('produkty:konta_lista')
    else:
        form = UserCreateForm()
    return render(request, 'produkty/konto_form.html', {'form': form, 'akcja': 'Dodaj'})

@staff_member_required
def konto_edytuj(request, user_id):
    konta_user = get_object_or_404(User, id=user_id)
    if request.method == 'POST':
        form = UserEditForm(request.POST, instance=konta_user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Zaktualizowano dane konta.')
            return redirect('produkty:konta_lista')
    else:
        form = UserEditForm(instance=konta_user)
    return render(request, 'produkty/konto_form.html', {'form': form, 'akcja': 'Edytuj'})

@staff_member_required
def konto_usun(request, user_id):
    konta_user = get_object_or_404(User, id=user_id)
    if request.method == 'POST':
        if konta_user != request.user and not konta_user.is_superuser:
            konta_user.delete()
            messages.success(request, 'Konto zostało usunięte.')
        else:
            messages.error(request, 'Nie możesz usunąć tego konta.')
        return redirect('produkty:konta_lista')
    return render(request, 'produkty/konto_usun.html', {'konta_user': konta_user})

from .forms import FunkcjaForm, PodpowiedzForm

@staff_member_required
def funkcja_dodaj(request):
    if request.method == 'POST':
        form = FunkcjaForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Funkcja została dodana.')
            return redirect('produkty:hub_wiedzy')
    else:
        form = FunkcjaForm()
    return render(request, 'produkty/formularz_wiedzy.html', {'form': form, 'akcja': 'Dodaj', 'typ': 'Funkcję'})

@staff_member_required
def funkcja_edytuj(request, funkcja_id):
    funkcja = get_object_or_404(Funkcja, id=funkcja_id)
    if request.method == 'POST':
        form = FunkcjaForm(request.POST, instance=funkcja)
        if form.is_valid():
            form.save()
            messages.success(request, 'Zaktualizowano funkcję.')
            return redirect('produkty:funkcja_szczegoly', funkcja_id=funkcja.id)
    else:
        form = FunkcjaForm(instance=funkcja)
    return render(request, 'produkty/formularz_wiedzy.html', {'form': form, 'akcja': 'Edytuj', 'typ': 'Funkcję'})

@staff_member_required
def podpowiedz_dodaj(request):
    if request.method == 'POST':
        form = PodpowiedzForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Podpowiedź została dodana.')
            return redirect('produkty:hub_wiedzy')
    else:
        form = PodpowiedzForm()
    return render(request, 'produkty/formularz_wiedzy.html', {'form': form, 'akcja': 'Dodaj', 'typ': 'Podpowiedź'})

@staff_member_required
def podpowiedz_edytuj(request, podpowiedz_id):
    podpowiedz = get_object_or_404(Podpowiedz, id=podpowiedz_id)
    if request.method == 'POST':
        form = PodpowiedzForm(request.POST, instance=podpowiedz)
        if form.is_valid():
            form.save()
            messages.success(request, 'Zaktualizowano podpowiedź.')
            return redirect('produkty:podpowiedz_szczegoly', podpowiedz_id=podpowiedz.id)
    else:
        form = PodpowiedzForm(instance=podpowiedz)
    return render(request, 'produkty/formularz_wiedzy.html', {'form': form, 'akcja': 'Edytuj', 'typ': 'Podpowiedź'})

import json
from collections import defaultdict
from django.http import JsonResponse

def ensure_default_subgroups():
    grupy_bez_subgrup = TopGroup.objects.filter(subgrupy__isnull=True)
    for grupa in grupy_bez_subgrup:
        TopSubGroup.objects.create(grupa=grupa, nazwa=grupa.nazwa, kolejnosc=0)

@login_required
def top_lista(request):
    ensure_default_subgroups()
    aktywne_grupy = TopGroup.objects.filter(aktywna=True).prefetch_related('subgrupy')
    wszystkie_modele = list(Produkt.objects.values_list('model', flat=True))
    
    # Przekazujemy wszystkie wpisy dla bieżącego użytkownika jako słownik {subgrupa_id: [wpisy]}
    wpisy_usera = TopListEntry.objects.filter(user=request.user).order_by('pozycja')
    wpisy_dict = defaultdict(list)
    for wpis in wpisy_usera:
        wpisy_dict[wpis.subgrupa_id].append({'pozycja': wpis.pozycja, 'model_tekst': wpis.model_tekst})
        
    return render(request, 'produkty/top_lista.html', {
        'aktywne_grupy': aktywne_grupy,
        'all_models_json': json.dumps(wszystkie_modele),
        'wpisy_dict': json.dumps(dict(wpisy_dict))
    })

@login_required
def top_lista_zapisz(request, subgrupa_id):
    if request.method == 'POST':
        subgrupa = get_object_or_404(TopSubGroup, id=subgrupa_id)
        # Parse JSON z requesta (spodziewamy się [{pozycja: 1, model: "X"}, ...])
        try:
            dane = json.loads(request.body)
            # Kasujemy stare wpisy dla tej podgrupy i usera
            TopListEntry.objects.filter(user=request.user, subgrupa=subgrupa).delete()
            # Tworzymy nowe
            nowe_wpisy = []
            for item in dane:
                model = item.get('model', '').strip().upper()
                pozycja = item.get('pozycja')
                if model and pozycja:
                    nowe_wpisy.append(TopListEntry(
                        user=request.user,
                        subgrupa=subgrupa,
                        pozycja=pozycja,
                        model_tekst=model
                    ))
            TopListEntry.objects.bulk_create(nowe_wpisy)
            return JsonResponse({'status': 'success'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'msg': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'msg': 'Invalid method'}, status=405)

@staff_member_required
def top_lista_manage(request):
    ensure_default_subgroups()
    grupy = TopGroup.objects.all().prefetch_related('subgrupy')
    return render(request, 'produkty/top_lista_manage.html', {'grupy': grupy})

@staff_member_required
def top_lista_group_add(request):
    if request.method == 'POST':
        nazwa = request.POST.get('nazwa', '').strip()
        kolejnosc = request.POST.get('kolejnosc', 0)
        if nazwa:
            TopGroup.objects.create(nazwa=nazwa, kolejnosc=kolejnosc)
            messages.success(request, 'Dodano grupę TOP.')
    return redirect('produkty:top_lista_manage')

@staff_member_required
def top_lista_subgroup_add(request):
    if request.method == 'POST':
        grupa_id = request.POST.get('grupa_id')
        nazwa = request.POST.get('nazwa', '').strip()
        kolejnosc = request.POST.get('kolejnosc', 0)
        if grupa_id and nazwa:
            grupa = get_object_or_404(TopGroup, id=grupa_id)
            TopSubGroup.objects.create(grupa=grupa, nazwa=nazwa, kolejnosc=kolejnosc)
            messages.success(request, 'Dodano podgrupę TOP.')
    return redirect('produkty:top_lista_manage')

@staff_member_required
def top_lista_delete_group(request, group_id):
    grupa = get_object_or_404(TopGroup, id=group_id)
    grupa.delete()
    messages.success(request, 'Usunięto grupę.')
    return redirect('produkty:top_lista_manage')

@staff_member_required
def top_lista_delete_subgroup(request, subgroup_id):
    subgrupa = get_object_or_404(TopSubGroup, id=subgroup_id)
    subgrupa.delete()
    messages.success(request, 'Usunięto podgrupę.')
    return redirect('produkty:top_lista_manage')


@login_required
def polecane_modele(request, year=None, month=None):
    # Determine the target month and year
    if year is None or month is None:
        today = timezone.now().date()
        year = today.year
        month = today.month

    from datetime import date
    import calendar
    
    current_month_start = date(year, month, 1)
    
    prev_month_end = current_month_start - timedelta(days=1)
    prev_year = prev_month_end.year
    prev_month = prev_month_end.month
    
    next_month_start = (current_month_start + timedelta(days=32)).replace(day=1)
    next_year = next_month_start.year
    next_month = next_month_start.month
    
    last_day_of_month = calendar.monthrange(year, month)[1]
    month_end_date = date(year, month, last_day_of_month)
    
    # 1. Fetch all products
    products = Produkt.objects.all()
    
    # 2. Historical sales for the target calendar month (across all years)
    sales_by_product_month = (
        Sprzedaz.objects.filter(data_sprzedazy__month=month)
        .values('produkt_id')
        .annotate(total_qty=Sum('liczba_sztuk'))
    )
    month_sales_dict = {item['produkt_id']: item['total_qty'] for item in sales_by_product_month}
    
    # 3. Total historical sales (all time)
    sales_by_product_total = (
        Sprzedaz.objects.values('produkt_id')
        .annotate(total_qty=Sum('liczba_sztuk'))
    )
    total_sales_dict = {item['produkt_id']: item['total_qty'] for item in sales_by_product_total}
    
    # 4. Active tasks in the target month
    active_tasks = Zadanie.objects.filter(
        data_start__lte=month_end_date,
        data_koniec__gte=current_month_start
    ).prefetch_related('produkty')
    
    # Map products to their active tasks
    product_to_tasks = defaultdict(list)
    for task in active_tasks:
        for p in task.produkty.all():
            product_to_tasks[p.id].append(task)
            
    # Group recommendations
    recommendations_raw = defaultdict(lambda: defaultdict(list))
    
    for p in products:
        prowizja_rate = p.stawka or Decimal('0.00')
        month_sales = month_sales_dict.get(p.id, 0)
        total_sales = total_sales_dict.get(p.id, 0)
        tasks = product_to_tasks.get(p.id, [])
        is_active_in_task = len(tasks) > 0
        
        # Scoring logic
        score = float(prowizja_rate) * 1.0 + float(month_sales) * 5.0 + float(total_sales) * 1.0 + (50.0 if is_active_in_task else 0.0)
        
        task_info = []
        for t in tasks:
            task_type = t.get_typ_display() if hasattr(t, 'get_typ_display') else t.typ
            task_info.append(f"{t.nazwa} ({task_type})")
            
        group_name = p.grupa_towarowa.strip().upper() if p.grupa_towarowa else 'INNE'
        brand_name = p.marka.strip().upper() if p.marka else 'INNE'
        
        recommendations_raw[group_name][brand_name].append({
            'id': p.id,
            'model': p.model,
            'stawka': prowizja_rate,
            'month_sales': month_sales,
            'total_sales': total_sales,
            'tasks': task_info,
            'score': score
        })
        
    # Sort and take top 4 for each group + brand
    grouped_recommendations = {}
    for group_name, brands in recommendations_raw.items():
        grouped_recommendations[group_name] = {}
        for brand_name, items in brands.items():
            sorted_items = sorted(
                items,
                key=lambda x: (-x['score'], -float(x['stawka']), -x['total_sales'], x['model'])
            )
            grouped_recommendations[group_name][brand_name] = sorted_items[:4]
            
    grouped_recommendations = {
        g: b for g, b in grouped_recommendations.items() if b
    }
    
    # Sort groups and brands alphabetically
    sorted_recommendations = sorted(
        [
            (
                g,
                sorted(
                    [(b, items) for b, items in brands.items()],
                    key=lambda x: x[0]
                )
            )
            for g, brands in grouped_recommendations.items()
        ],
        key=lambda x: x[0]
    )
    
    # Translate month name to Polish
    months_pl = {
        1: 'Styczeń', 2: 'Luty', 3: 'Marzec', 4: 'Kwiecień', 5: 'Maj', 6: 'Czerwiec',
        7: 'Lipiec', 8: 'Sierpień', 9: 'Wrzesień', 10: 'Październik', 11: 'Listopad', 12: 'Grudzień'
    }
    month_name_pl = f"{months_pl.get(month, '')} {year}"
    
    context = {
        'recommendations': sorted_recommendations,
        'year': year,
        'month': month,
        'month_name': month_name_pl,
        'prev_year': prev_year,
        'prev_month': prev_month,
        'next_year': next_year,
        'next_month': next_month,
    }
    return render(request, 'produkty/polecane_modele.html', context)
